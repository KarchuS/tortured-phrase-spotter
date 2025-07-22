# Модуль для работы со списком фраз

def load_phrases_from_xlsx(xlsx_path):
    """
    Загружает список фраз из xlsx-файла (экспортированного из Numbers).
    Ожидается, что фразы находятся в первом столбце, первая строка — заголовок.
    """
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active
    phrases = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # пропускаем заголовок
        phrase = row[0]
        if phrase and isinstance(phrase, str):
            phrases.append(phrase.strip())
    return phrases


def load_phrases_and_recommendations_from_csv(csv_path):
    """
    Загружает фразы для поиска и рекомендации из csv-файла.
    Возвращает (список фраз, словарь рекомендаций).
    """
    import csv
    phrases = []
    recommendations = {}
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader)  # пропускаем заголовок
        for row in reader:
            if len(row) < 2:
                continue
            phrase = row[0].strip(' "')
            rec = row[1].strip(' "')
            if phrase:
                phrases.append(phrase)
                recommendations[phrase] = rec
    return phrases, recommendations 